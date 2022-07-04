from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border,Alignment

def writeFile(pjLst):
    # 打开模板文件，准备写入报告
    wb = load_workbook('./conf/template.xlsx')
    ws = wb["Sheet1"]

    count=2
    align = Alignment(vertical='center', wrap_text=True)
    Fill=PatternFill(start_color="ffff00", fill_type="solid")
    for pj in pjLst:
        ws.row_dimensions[count].height = 30
        ws.cell(count,1,pj["contractNum"]).alignment = align
        ws.cell(count,2,pj["supplier"]).alignment = align
        ws.cell(count, 3, pj["custormer"]).alignment = align  # 客户
        ws.cell(count, 4, pj["product"]).alignment = align
        ws.cell(count, 5, pj["purchaseNum"]).alignment = align
        ws.cell(count, 6, pj["purchasePrice"]).alignment = align
        ws.cell(count, 7, pj["purchaseAmt"]).alignment = align  # 采购金额
        ws.cell(count, 8, pj["cost"]).alignment = align  # 成本
        ws.cell(count, 9, pj["paidAmt"]).alignment = align  # 累计付款金额
        if pj["purchaseAmt"]!=pj["paidAmt"]:
            ws.cell(count, 9).fill = Fill
        details=[]
        num=1
        for d in pj["details"]:
            if d["flag"]==0:
                details.append("{}. {} {}".format(num,d["recPayDate"],d["amt"]))
                num=num+1
        dd="\n".join(details)[:-1]
        ws.cell(count, 10,dd).alignment = align  # 付款明细
        ws.cell(count, 11, pj["inputVat"]).alignment = align  # 进项税额
        ws.cell(count, 12, pj["saleNum"]).alignment = align
        ws.cell(count, 13, pj["salePrice"]).alignment = align
        ws.cell(count,14, pj["saleAmt"]).alignment = align  # 销售金额
        ws.cell(count, 15, pj["inputAmt"]).alignment = align  # 收入
        ws.cell(count, 16, pj["receivedAmt"]).alignment = align  # 累计收款金额
        if pj["saleAmt"]!=pj["receivedAmt"]:
            ws.cell(count, 16).fill = Fill
        details = []
        num = 1
        for d in pj["details"]:
            if d["flag"]==1:
                details.append("{}. {} {}".format(num,d["recPayDate"],d["amt"]))
                num=num+1
        dd="\n".join(details)[:-1]
        ws.cell(count, 17,dd).alignment = align  #收款明细
        ws.cell(count, 18, pj["outputVat"]).alignment = align  # 销项税额
        ws.cell(count, 19, pj["grossPft"]).alignment = align  # 毛利
        ws.cell(count, 20, pj["addTax"]).alignment = align  # 增值税
        ws.cell(count, 21, pj["surTax"]).alignment = align  # 附加税
        ws.cell(count, 22, pj["stampTax"]).alignment = align  # 印花税
        ws.cell(count, 23, pj["nt"]).alignment = align  # 净利润
        isInputFapiao="否" if pj["isInputFapiao"]==0 else "是"
        ws.cell(count, 24,isInputFapiao).alignment = align  # 是否收票
        if isInputFapiao == "是":
            inputFapiaoDate=pj["inputFapiaoDate"]
            if inputFapiaoDate=="2000-01-01":
                ws.cell(count, 25,"未收票").fill = Fill
                ws.cell(count, 26).fill = Fill

            else:
                ws.cell(count, 25, pj["inputFapiaoDate"]).alignment = align  # 收票日期
                isSend="否" if pj["isSend"]==0 else "是"
                ws.cell(count, 26, isSend).alignment = align  # 是否寄出
                if isSend == "否":
                    ws.cell(count, 26).fill = Fill
        isMakeFapiao="否" if pj["isMakeFapiao"]==0 else "是"
        ws.cell(count, 27, isMakeFapiao).alignment = align  # 开票
        if isMakeFapiao=="是":
            if pj["makeFapiaoDate"]=="2000-01-01":
                ws.cell(count, 28,"未收票").fill = Fill
            else:
                ws.cell(count, 28, pj["makeFapiaoDate"]).alignment = align  # 开票日期
        ws.cell(count, 29, pj["pjStatus"]).alignment = align    # 项目状态
        ws.cell(count, 30, pj["reverse"]).alignment = align # 备注
        count=count+1
    try:
        wb.save("./conf/台账.xlsx")
        res=""
    except:
        res="请先关闭台账文件"
    return res
